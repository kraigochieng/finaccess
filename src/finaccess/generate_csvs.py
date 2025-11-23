import openpyxl
import csv

# Open the Excel workbook
workbook = openpyxl.load_workbook("2024_Finaccess_Publicdata.xlsx")

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    # Open the sheet
    sheet = workbook[sheet_name]
    # Create a CSV file
    with open(f"{sheet_name}.csv", "w", newline="") as csvfile:
        writer = csv.writer(csvfile)
        # Iterate through rows in the sheet and write to CSV
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
