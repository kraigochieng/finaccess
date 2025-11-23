# optimized_excel_to_parquet.py

import pandas as pd
from openpyxl import load_workbook
import os
import sys
import pyarrow as pa
import pyarrow.parquet as pq


def convert_excel_to_parquet_optimized(
    excel_file_path: str,
    parquet_file_path: str,
    chunk_size: int = 50_000,
    compression: str = "zstd",  # zstd for balance; "snappy" for speed
):
    if os.path.exists(parquet_file_path):
        os.remove(parquet_file_path)
        print(f"Removed existing file: {parquet_file_path}")

    print(f"Streaming '{excel_file_path}' in chunks of {chunk_size:,} rows")
    print(f"Output → {parquet_file_path} (compression: {compression})")

    # Load workbook in read-only mode (minimal memory)
    workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
    sheet = workbook.active

    # Get header and total row estimate (optional, for progress)
    rows_iterator = sheet.iter_rows(values_only=True)
    header = next(rows_iterator)  # First row = header
    print(f"Detected {len(header)} columns: {header[:5]}...")  # Preview

    # Clean header: make valid Parquet column names
    header = [
        f"col_{i}"
        if col is None
        else str(col).strip().replace(" ", "_").replace("\n", "_")
        for i, col in enumerate(header)
    ]

    writer = None
    total_rows = 0
    chunk_count = 0
    chunk_data = []

    try:
        for row in rows_iterator:
            # Clean row: None → empty string (Parquet-friendly)
            cleaned_row = tuple(str(cell) if cell is not None else "" for cell in row)
            chunk_data.append(cleaned_row)

            if len(chunk_data) >= chunk_size:
                # Build DataFrame and infer dtypes (better than all-str)
                df_chunk = pd.DataFrame(chunk_data, columns=header)
                # Optional: smarter dtypes (uncomment if you want numeric detection; costs a bit more RAM)
                # for col in df_chunk.select_dtypes(include='object').columns:
                #     df_chunk[col] = pd.to_numeric(df_chunk[col], errors='coerce').fillna(df_chunk[col])

                # Fill NaNs (from dtype inference if used)
                df_chunk = df_chunk.fillna("")

                # To PyArrow Table
                table = pa.Table.from_pandas(df_chunk, preserve_index=False)

                if writer is None:
                    # First chunk: init writer with schema
                    writer = pq.ParquetWriter(
                        parquet_file_path,
                        schema=table.schema,
                        compression=compression,
                        use_dictionary=True,  # Compresses repeated strings well
                        write_statistics=True,
                    )

                writer.write_table(table)
                total_rows += len(chunk_data)
                chunk_count += 1

                print(
                    f"Chunk {chunk_count:,}: {len(chunk_data):,} rows → {df_chunk.shape}"
                )

                chunk_data = []  # Reset

        # Final chunk
        if chunk_data:
            df_chunk = pd.DataFrame(chunk_data, columns=header)
            df_chunk = df_chunk.fillna("")
            table = pa.Table.from_pandas(df_chunk, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(
                    parquet_file_path, table.schema, compression=compression
                )
            writer.write_table(table)
            total_rows += len(chunk_data)
            print(f"Final chunk: {len(chunk_data):,} rows")

    except Exception as e:
        print(f"Error during processing: {e}")
        sys.exit(1)
    finally:
        if writer:
            writer.close()
        workbook.close()

    file_size_gb = os.path.getsize(parquet_file_path) / 1e9
    print(
        f"\nSUCCESS! {total_rows:,} rows → {parquet_file_path} ({file_size_gb:.2f} GB)"
    )


if __name__ == "__main__":
    convert_excel_to_parquet_optimized(
        excel_file_path="2024_Finaccess_Publicdata.xlsx",
        parquet_file_path="finaccess_2024_optimized.parquet",
        chunk_size=1_000,  # Start here; bump to 100K if you have 16GB+ RAM
        compression="zstd",
    )
