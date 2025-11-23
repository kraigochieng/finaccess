# 1_excel_to_csv_streaming.py
from openpyxl import load_workbook
import csv
import os
import sys
from pathlib import Path

def excel_to_csv_streaming(
    excel_path="2024_Finaccess_Publicdata.xlsx",
    csv_path="2024_Finaccess_Publicdata_from_script.csv",
    chunk_size=50_000
):
    print(f"Streaming {excel_path} → {csv_path} (low-memory mode)")

    if os.path.exists(csv_path):
        os.remove(csv_path)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    # Clean header once
    clean_header = [
        str(col).strip().replace(",", "_").replace("\n", " ").replace('"', "'")
        if col is not None else f"empty_col_{i}"
        for i, col in enumerate(header)
    ]

    total_rows = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(clean_header)

        batch = []
        for row in rows:
            cleaned_row = [
                "" if cell is None else str(cell).replace(",", " ").replace("\n", " ").replace('"', "'")
                for cell in row
            ]
            batch.append(cleaned_row)

            if len(batch) >= chunk_size:
                writer.writerows(batch)
                total_rows += len(batch)
                print(f"Written {total_rows:,} rows...")
                batch = []

        # Final batch
        if batch:
            writer.writerows(batch)
            total_rows += len(batch)
            print(f"Final batch: {len(batch):,} rows")

    wb.close()
    file_size = Path(csv_path).stat().st_size / 1e9
    print(f"DONE! {total_rows:,} rows → {csv_path} ({file_size:.2f} GB)")

if __name__ == "__main__":
    excel_to_csv_streaming(chunk_size=100_000)  # 100k = fastest on 16GB RAM